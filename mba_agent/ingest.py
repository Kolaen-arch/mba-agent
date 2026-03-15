"""
PDF ingestion: extract text, clean, chunk, and prepare for embedding.
Handles English and Danish PDFs with proper character handling.
Filters out reference sections, headers/footers, figure captions, and noise.
"""

import os
import re
from dataclasses import dataclass, field
from pathlib import Path

import fitz  # pymupdf


@dataclass
class Chunk:
    """A text chunk with source metadata."""
    text: str
    source_file: str
    page_start: int
    page_end: int
    chunk_index: int
    metadata: dict = field(default_factory=dict)
    label: str = ""  # User-defined tag, e.g. "my paper, first semester"
    section_header: str = ""  # Detected heading, e.g. "Literature Review"
    doc_type: str = ""  # Classified type: own_work, textbook, article, case

    @property
    def source_tag(self) -> str:
        prefix = f"[{self.label}] " if self.label else ""
        sec = f" | Section: {self.section_header}" if self.section_header else ""
        if self.page_start == self.page_end:
            return f"{prefix}[SOURCE: {self.source_file}, page {self.page_start}{sec}]"
        return f"{prefix}[SOURCE: {self.source_file}, pages {self.page_start}-{self.page_end}{sec}]"

    @property
    def embedding_text(self) -> str:
        """Text with contextual header prepended for embedding."""
        header_parts = [f"Document: {self.source_file}"]
        if self.section_header:
            header_parts.append(f"Section: {self.section_header}")
        header = " | ".join(header_parts)
        return f"[{header}]\n{self.text}"

    def to_context_str(self) -> str:
        return f"{self.source_tag}\n{self.text}"


# ── Cleaning ──

# Patterns for headers/footers (page numbers, journal names, DOIs)
HEADER_FOOTER_PATTERNS = [
    r'^\d+\s*$',                              # Bare page numbers
    r'^Page \d+ of \d+',                      # "Page X of Y"
    r'^\d+\s+[A-Z][a-z]+ et al\.',           # "42 Smith et al."
    r'^(Downloaded|Available) (from|at) http', # Download notices
    r'^https?://\S+$',                        # Bare URLs
    r'^DOI:?\s*\d',                           # DOI lines
    r'^\d{4}\s+\w+\s+\d+\s+\w+$',           # Date headers
    r'^©\s*\d{4}',                            # Copyright lines
    r'^ISSN\s',                               # ISSN
    r'^\[\d+\]$',                             # Bare reference numbers
    r'^Vol\.?\s*\d',                          # Volume headers
    r'^Journal of\s',                         # Journal name headers
    r'^This article\s',                       # Standard boilerplate
]
_header_re = [re.compile(p, re.IGNORECASE) for p in HEADER_FOOTER_PATTERNS]

# Patterns that indicate a references section has begun
REFS_START_PATTERNS = [
    r'^References?\s*$',
    r'^Bibliography\s*$',
    r'^Works Cited\s*$',
    r'^Litteratur(?:liste|fortegnelse)?\s*$',  # Danish
    r'^Kilder\s*$',                             # Danish: "Sources"
    r'^Referanser?\s*$',                        # Norwegian (close to Danish)
    r'^REFERENCES?\s*$',
]
_refs_start_re = [re.compile(p) for p in REFS_START_PATTERNS]

# Figure/table caption patterns
CAPTION_PATTERNS = [
    r'^(?:Figure|Fig\.?|Table|Tabel|Figur)\s+\d',  # "Figure 1", "Tabel 2"
    r'^(?:Source|Kilde|Note|Anm):',                 # Source/Note lines
]
_caption_re = [re.compile(p, re.IGNORECASE) for p in CAPTION_PATTERNS]


def _is_header_footer(line: str) -> bool:
    line = line.strip()
    if len(line) < 3:
        return True
    return any(r.match(line) for r in _header_re)


def _is_refs_start(line: str) -> bool:
    return any(r.match(line.strip()) for r in _refs_start_re)


def _is_caption(line: str) -> bool:
    return any(r.match(line.strip()) for r in _caption_re)


def _fix_encoding(text: str) -> str:
    """Fix common encoding artifacts in Danish/Swedish text."""
    fixes = {
        'Ã¦': 'æ', 'Ã¸': 'ø', 'Ã¥': 'å',
        'Ã†': 'Æ', 'Ã˜': 'Ø', 'Ã…': 'Å',
        'Ã¤': 'ä', 'Ã¶': 'ö', 'Ã¼': 'ü',
        'Ã„': 'Ä', 'Ã–': 'Ö', 'Ãœ': 'Ü',
        'â€™': "'", 'â€œ': '"', 'â€\x9d': '"',
        'â€"': '—', 'â€"': '–',
        'ï¬': 'fi', 'ï¬‚': 'fl',  # Ligature artifacts
        '\x00': '', '\ufffd': '',  # Null bytes, replacement chars
    }
    for bad, good in fixes.items():
        text = text.replace(bad, good)
    return text


def _clean_page_text(text: str) -> str:
    """Clean a single page's extracted text."""
    text = _fix_encoding(text)

    lines = text.split('\n')
    cleaned = []

    for line in lines:
        stripped = line.strip()

        # Skip empty/tiny lines
        if len(stripped) < 4:
            continue

        # Skip headers/footers
        if _is_header_footer(stripped):
            continue

        # Skip figure/table captions (they add noise, not substance)
        if _is_caption(stripped):
            continue

        # Fix hyphenation at line breaks (common in PDFs)
        if cleaned and cleaned[-1].endswith('-'):
            # Join hyphenated word
            prev = cleaned[-1][:-1]
            cleaned[-1] = prev + stripped
        else:
            cleaned.append(stripped)

    return '\n'.join(cleaned)


def _detect_headings_from_dict(page) -> list[tuple[str, float]]:
    """Detect headings by font size from page dict blocks.
    Returns list of (heading_text, font_size) found on the page."""
    try:
        blocks = page.get_text("dict")["blocks"]
    except Exception:
        return []

    # Collect all text spans with their font sizes
    spans_info = []
    for block in blocks:
        if block.get("type") != 0:  # text blocks only
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                text = span.get("text", "").strip()
                size = span.get("size", 0)
                if text and size > 0:
                    spans_info.append((text, size))

    if len(spans_info) < 3:
        return []

    # Median font size = body text size
    sizes = sorted(s for _, s in spans_info)
    median_size = sizes[len(sizes) // 2]

    # Headings: significantly larger than body, reasonable length
    headings = []
    for text, size in spans_info:
        if size > median_size * 1.15 and 3 < len(text) < 200:
            # Skip lines that look like page numbers or artifacts
            if re.match(r'^\d+\s*$', text):
                continue
            headings.append((text, size))

    return headings


def extract_pdf_text(
    pdf_path: str,
    strip_references: bool = True,
    detect_headings: bool = True,
) -> list[tuple[int, str, str]]:
    """
    Extract and clean text from PDF.
    Returns list of (page_number, text, current_heading) tuples.
    Stops at the References section if strip_references=True.
    """
    doc = fitz.open(pdf_path)
    pages = []
    refs_started = False
    current_heading = ""

    for page_num in range(len(doc)):
        page = doc[page_num]
        raw_text = page.get_text("text")

        # Detect headings from font-size analysis
        if detect_headings:
            headings = _detect_headings_from_dict(page)
            if headings:
                # Use the last heading on the page (most likely the active section)
                current_heading = headings[-1][0]

        # Clean
        text = _clean_page_text(raw_text)

        # Check if references section starts on this page
        if strip_references and not refs_started:
            for line in text.split('\n'):
                if _is_refs_start(line):
                    idx = text.find(line)
                    if idx > 0:
                        text = text[:idx].strip()
                    else:
                        text = ""
                    refs_started = True
                    break

        if refs_started and not text:
            continue

        # Normalize whitespace
        text = re.sub(r'\n{3,}', '\n\n', text)
        text = re.sub(r'[ \t]{2,}', ' ', text)
        text = text.strip()

        if text and len(text) > 50:
            pages.append((page_num + 1, text, current_heading))

    doc.close()
    return pages


def extract_pdf_metadata(pdf_path: str) -> dict:
    """Extract metadata from PDF."""
    doc = fitz.open(pdf_path)
    meta = doc.metadata or {}
    page_count = len(doc)
    doc.close()

    return {
        "title": _fix_encoding(meta.get("title", "")),
        "author": _fix_encoding(meta.get("author", "")),
        "subject": _fix_encoding(meta.get("subject", "")),
        "filename": os.path.basename(pdf_path),
        "page_count": page_count,
    }


def _is_substantive_paragraph(text: str) -> bool:
    """Filter out non-substantive text blocks."""
    text = text.strip()
    if len(text) < 40:
        return False

    # Skip lines that are mostly numbers or symbols
    alpha_ratio = sum(c.isalpha() for c in text) / max(len(text), 1)
    if alpha_ratio < 0.5:
        return False

    # Skip very short "paragraphs" that are likely artifacts
    words = text.split()
    if len(words) < 8:
        return False

    return True


def chunk_pages(
    pages: list[tuple[int, str]] | list[tuple[int, str, str]],
    source_file: str,
    chunk_size: int = 1500,
    chunk_overlap: int = 200,
) -> list[Chunk]:
    """
    Chunk extracted pages into overlapping segments.
    chunk_size and chunk_overlap are in characters.
    Pages can be (page_num, text) or (page_num, text, section_header).
    """
    chunks = []
    segments = []  # (page_num, para_text, section_header)

    for page in pages:
        page_num = page[0]
        text = page[1]
        heading = page[2] if len(page) > 2 else ""
        paragraphs = text.split('\n\n')
        for para in paragraphs:
            para = para.strip()
            if _is_substantive_paragraph(para):
                segments.append((page_num, para, heading))

    if not segments:
        return chunks

    current_text = ""
    current_page_start = segments[0][0]
    current_page_end = current_page_start
    current_heading = segments[0][2]
    chunk_idx = 0
    char_limit = chunk_size * 4  # Convert to chars

    for page_num, para, heading in segments:
        if heading:
            current_heading = heading
        test_text = current_text + "\n\n" + para if current_text else para

        if len(test_text) > char_limit and current_text:
            chunks.append(Chunk(
                text=current_text.strip(),
                source_file=source_file,
                page_start=current_page_start,
                page_end=current_page_end,
                chunk_index=chunk_idx,
                section_header=current_heading,
            ))
            chunk_idx += 1

            overlap_text = current_text[-(chunk_overlap * 4):]
            current_text = overlap_text + "\n\n" + para
            current_page_start = current_page_end
            current_page_end = page_num
        else:
            current_text = test_text
            current_page_end = page_num

    if current_text.strip():
        chunks.append(Chunk(
            text=current_text.strip(),
            source_file=source_file,
            page_start=current_page_start,
            page_end=current_page_end,
            chunk_index=chunk_idx,
            section_header=current_heading,
        ))

    return chunks


def extract_xlsx_chunks(xlsx_path: str, chunk_size_rows: int = 50) -> list[Chunk]:
    """
    Extract tabular data from Excel files.
    Per-sheet: summary chunk + data chunks of chunk_size_rows rows each.
    """
    import openpyxl
    filename = os.path.basename(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    chunks = []
    chunk_idx = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(c) if c is not None else "" for c in rows[0]]
        data_rows = rows[1:]

        # Summary chunk: sheet name + headers + first 5 rows
        preview_lines = [" | ".join(h for h in headers if h)]
        for row in data_rows[:5]:
            preview_lines.append(" | ".join(str(c) if c is not None else "" for c in row))
        summary_text = (
            f"Excel sheet: {sheet_name}\n"
            f"Columns: {', '.join(h for h in headers if h)}\n"
            f"Rows: {len(data_rows)}\n\n"
            f"Preview:\n" + "\n".join(preview_lines)
        )
        chunks.append(Chunk(
            text=summary_text, source_file=filename,
            page_start=0, page_end=0, chunk_index=chunk_idx,
            metadata={"sheet_name": sheet_name, "file_type": "xlsx", "is_summary": True},
        ))
        chunk_idx += 1

        # Data chunks
        for i in range(0, len(data_rows), chunk_size_rows):
            batch = data_rows[i:i + chunk_size_rows]
            lines = [" | ".join(headers)]
            for row in batch:
                lines.append(" | ".join(str(c) if c is not None else "" for c in row))
            text = f"[Sheet: {sheet_name}, rows {i+1}-{min(i+chunk_size_rows, len(data_rows))}]\n" + "\n".join(lines)
            if len(text.strip()) < 40:
                continue
            chunks.append(Chunk(
                text=text, source_file=filename,
                page_start=i, page_end=min(i + chunk_size_rows, len(data_rows)),
                chunk_index=chunk_idx,
                metadata={"sheet_name": sheet_name, "file_type": "xlsx"},
            ))
            chunk_idx += 1

    wb.close()
    return chunks


def extract_csv_chunks(csv_path: str, chunk_size_rows: int = 100) -> list[Chunk]:
    """Extract tabular data from CSV files."""
    import csv as csv_module
    filename = os.path.basename(csv_path)
    chunks = []
    chunk_idx = 0

    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv_module.reader(f)
        rows = list(reader)

    if not rows:
        return chunks

    headers = rows[0]
    data_rows = rows[1:]

    # Summary chunk
    summary_text = (
        f"CSV file: {filename}\n"
        f"Columns: {', '.join(headers)}\n"
        f"Rows: {len(data_rows)}"
    )
    chunks.append(Chunk(
        text=summary_text, source_file=filename,
        page_start=0, page_end=0, chunk_index=chunk_idx,
        metadata={"file_type": "csv", "is_summary": True},
    ))
    chunk_idx += 1

    for i in range(0, len(data_rows), chunk_size_rows):
        batch = data_rows[i:i + chunk_size_rows]
        lines = [" | ".join(headers)]
        lines.extend(" | ".join(row) for row in batch)
        text = f"[Rows {i+1}-{min(i+chunk_size_rows, len(data_rows))}]\n" + "\n".join(lines)
        chunks.append(Chunk(
            text=text, source_file=filename,
            page_start=i, page_end=min(i + chunk_size_rows, len(data_rows)),
            chunk_index=chunk_idx,
            metadata={"file_type": "csv"},
        ))
        chunk_idx += 1

    return chunks


def extract_docx_chunks(
    docx_path: str,
    chunk_size: int = 1500,
    chunk_overlap: int = 200,
) -> list[Chunk]:
    """Extract content from DOCX with heading detection via paragraph styles."""
    from docx import Document as DocxDocument
    filename = os.path.basename(docx_path)
    doc = DocxDocument(docx_path)

    pages = []  # (page_num, text, heading)
    current_page = 1
    current_parts = []
    current_heading = ""
    char_count = 0

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        # Detect headings from Word styles (Heading 1, Heading 2, etc.)
        style_name = para.style.name if para.style else ""
        if style_name.startswith("Heading"):
            current_heading = text

        current_parts.append(text)
        char_count += len(text)
        if char_count >= 2400:  # ~1 page
            pages.append((current_page, "\n\n".join(current_parts), current_heading))
            current_page += 1
            current_parts = []
            char_count = 0

    if current_parts:
        pages.append((current_page, "\n\n".join(current_parts), current_heading))

    return chunk_pages(pages, filename, chunk_size=chunk_size, chunk_overlap=chunk_overlap)


def extract_text_chunks(
    text_path: str,
    chunk_size: int = 1500,
    chunk_overlap: int = 200,
) -> list[Chunk]:
    """Extract content from plain text or markdown files."""
    filename = os.path.basename(text_path)
    with open(text_path, encoding='utf-8', errors='replace') as f:
        content = f.read()

    paragraphs = re.split(r'\n\n+', content)
    pages = []
    current_page = 1
    current_batch = []
    char_count = 0

    for para in paragraphs:
        para = para.strip()
        if not para or len(para) < 20:
            continue
        current_batch.append(para)
        char_count += len(para)
        if char_count >= 2400:
            pages.append((current_page, "\n\n".join(current_batch)))
            current_page += 1
            current_batch = []
            char_count = 0

    if current_batch:
        pages.append((current_page, "\n\n".join(current_batch)))

    return chunk_pages(pages, filename, chunk_size=chunk_size, chunk_overlap=chunk_overlap)


def extract_full_text(pdf_path: str, strip_references: bool = True) -> str:
    """Extract cleaned text as a single string. For Gemini full-context path."""
    pages = extract_pdf_text(pdf_path, strip_references=strip_references, detect_headings=False)
    return "\n\n".join(text for _, text, *_ in pages)


def classify_doc_type(filename: str, text_sample: str, label: str = "") -> str:
    """Heuristic document type classification for retrieval biasing."""
    fn = filename.lower()
    lab = label.lower()

    # User's own work
    if any(k in lab for k in ("my paper", "min opgave", "draft", "udkast", "own")):
        return "own_work"
    if any(k in fn for k in ("draft", "udkast", "opgave", "assignment")):
        return "own_work"

    # Textbook detection
    if "chapter" in fn or "kapitel" in fn:
        return "textbook"

    # Check text sample for article vs textbook vs case signals
    sample = text_sample[:2000].lower()
    if "abstract" in sample and ("doi" in sample or "et al" in sample):
        return "article"
    if "case study" in sample[:500] or "case" in fn:
        return "case"
    if any(w in sample for w in ("learning objectives", "key terms", "chapter summary")):
        return "textbook"

    return "article"  # Default


def ingest_files(
    file_paths: list[str],
    label: str = "",
    chunk_size: int = 1500,
    chunk_overlap: int = 200,
    strip_references: bool = True,
) -> tuple[list[Chunk], list[dict]]:
    """Ingest specific files with an optional label tag."""
    all_chunks = []
    all_metadata = []
    skipped = 0

    for fp in file_paths:
        file_path = Path(fp)
        filename = file_path.name
        ext = file_path.suffix.lower()

        try:
            if ext == ".pdf":
                pages = extract_pdf_text(str(file_path), strip_references=strip_references)
                if not pages:
                    skipped += 1
                    continue
                meta = extract_pdf_metadata(str(file_path))
                all_metadata.append(meta)
                chunks = chunk_pages(pages, filename, chunk_size=chunk_size, chunk_overlap=chunk_overlap)
            elif ext == ".xlsx":
                chunks = extract_xlsx_chunks(str(file_path))
            elif ext == ".csv":
                chunks = extract_csv_chunks(str(file_path))
            elif ext == ".docx":
                chunks = extract_docx_chunks(str(file_path), chunk_size=chunk_size, chunk_overlap=chunk_overlap)
            elif ext in (".txt", ".md"):
                chunks = extract_text_chunks(str(file_path), chunk_size=chunk_size, chunk_overlap=chunk_overlap)
            else:
                continue

            # Classify document type and apply label
            sample = chunks[0].text if chunks else ""
            dtype = classify_doc_type(filename, sample, label)
            for c in chunks:
                if label:
                    c.label = label
                c.doc_type = dtype
            all_chunks.extend(chunks)

        except Exception as e:
            print(f"  x Error processing {filename}: {e}")
            skipped += 1

    return all_chunks, all_metadata


def ingest_directory(
    papers_dir: str,
    chunk_size: int = 1500,
    chunk_overlap: int = 200,
    strip_references: bool = True,
    label: str = "",
) -> tuple[list[Chunk], list[dict]]:
    """Ingest all supported files from a directory (PDF, XLSX, CSV, DOCX, TXT, MD)."""
    papers_path = Path(papers_dir)

    # Gather all supported files
    all_files = sorted(
        list(papers_path.glob("**/*.pdf")) +
        list(papers_path.glob("**/*.xlsx")) +
        list(papers_path.glob("**/*.csv")) +
        list(papers_path.glob("**/*.docx")) +
        list(papers_path.glob("**/*.txt")) +
        list(papers_path.glob("**/*.md"))
    )

    all_chunks = []
    all_metadata = []
    skipped = 0

    for file_path in all_files:
        filename = file_path.name
        ext = file_path.suffix.lower()

        try:
            if ext == ".pdf":
                pages = extract_pdf_text(str(file_path), strip_references=strip_references)
                if not pages:
                    print(f"  ! Skipped (no extractable text): {filename}")
                    skipped += 1
                    continue
                meta = extract_pdf_metadata(str(file_path))
                all_metadata.append(meta)
                chunks = chunk_pages(pages, filename, chunk_size=chunk_size, chunk_overlap=chunk_overlap)
                print(f"  + {filename}: {meta['page_count']}p -> {len(chunks)} chunks")

            elif ext == ".xlsx":
                chunks = extract_xlsx_chunks(str(file_path))
                print(f"  + {filename}: {len(chunks)} chunks (Excel)")

            elif ext == ".csv":
                chunks = extract_csv_chunks(str(file_path))
                print(f"  + {filename}: {len(chunks)} chunks (CSV)")

            elif ext == ".docx":
                chunks = extract_docx_chunks(str(file_path), chunk_size=chunk_size, chunk_overlap=chunk_overlap)
                print(f"  + {filename}: {len(chunks)} chunks (DOCX)")

            elif ext in (".txt", ".md"):
                chunks = extract_text_chunks(str(file_path), chunk_size=chunk_size, chunk_overlap=chunk_overlap)
                print(f"  + {filename}: {len(chunks)} chunks (text)")

            else:
                continue

            # Classify document type and apply label
            sample = chunks[0].text if chunks else ""
            dtype = classify_doc_type(filename, sample, label)
            for c in chunks:
                if label:
                    c.label = label
                c.doc_type = dtype
            all_chunks.extend(chunks)

        except Exception as e:
            print(f"  x Error processing {filename}: {e}")
            skipped += 1

    if skipped:
        print(f"\n  ! {skipped} files skipped.")

    return all_chunks, all_metadata
